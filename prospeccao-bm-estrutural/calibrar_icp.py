#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calibra o ICP cruzando a lista de clientes atuais da BM Estrutural com o
indice regional da Receita Federal.

Responde tres perguntas que decidem o recorte da base:
  1. Que CNAE, capital, porte e cidade os clientes que JA COMPRAM realmente tem?
  2. Quantos deles seriam DESCARTADOS pelos filtros atuais do config.py?
  3. Que faturamento cada perfil gera (se a planilha trouxer valor de venda)?

Pre-requisito:
    python gerar_base.py --indice-amplo

Uso:
    python calibrar_icp.py clientes_ricardo.xlsx
    python calibrar_icp.py clientes.csv --indice saida/indice_regional_2026-08.csv

A planilha do cliente pode vir suja. O script detecta as colunas sozinho —
aceita CNPJ formatado ou nao, e casa por nome quando nao houver CNPJ.
"""

import argparse
import csv
import glob
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from difflib import SequenceMatcher

import config as C

DIR = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(DIR, "saida")

# Sufixos societarios e ruido que atrapalham o casamento por nome.
RUIDO = {
    "LTDA", "ME", "EPP", "SA", "S/A", "EIRELI", "MEI", "CIA", "E", "DE", "DA",
    "DO", "DAS", "DOS", "EM", "COM", "&", "-", "SS", "SIMPLES", "LIMITADA",
    "SOCIEDADE", "EMPRESA", "COMERCIO", "SERVICOS", "SERVICO", "INDUSTRIA",
    "PARTICIPACOES", "EMPREENDIMENTOS", "IMOBILIARIOS", "GERAIS",
}

# Como identificar cada coluna na planilha do cliente (busca por substring).
PISTAS = {
    "cnpj":      ["cnpj", "cpf/cnpj", "documento", "doc"],
    "nome":      ["razao", "cliente", "nome", "empresa", "fantasia", "construtora"],
    "valor":     ["valor", "total", "faturamento", "venda", "receita", "preco",
                  "ticket", "orcamento", "vlr"],
    "municipio": ["municipio", "cidade", "local"],
}


# ---------------------------------------------------------------------------
# Normalizacao e casamento
# ---------------------------------------------------------------------------

def norm(s):
    """Maiuscula, sem acento, so alfanumerico e espaco."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", re.sub(r"[^A-Za-z0-9 ]", " ", s)).upper().strip()


def tokens(s):
    """Tokens significativos: sem sufixo societario, sem token de 1 letra."""
    return [t for t in norm(s).split() if t not in RUIDO and len(t) > 1]


def chave(s):
    return " ".join(tokens(s))


def so_digitos(s):
    return re.sub(r"\D", "", str(s or ""))


def similaridade(a, b):
    """0-1 combinando cobertura de tokens e semelhanca literal."""
    ta, tb = set(tokens(a)), set(tokens(b))
    if not ta or not tb:
        return 0.0
    cobertura = len(ta & tb) / min(len(ta), len(tb))
    literal = SequenceMatcher(None, chave(a), chave(b)).ratio()
    return 0.65 * cobertura + 0.35 * literal


# ---------------------------------------------------------------------------
# Leitura da planilha do cliente
# ---------------------------------------------------------------------------

def ler_planilha(caminho):
    """Le CSV ou XLSX e devolve lista de dicts. Detecta separador e encoding."""
    ext = os.path.splitext(caminho)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Para ler .xlsx: pip install openpyxl "
                     "(ou salve a planilha como CSV)")
        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        linhas = [[("" if c is None else c) for c in r]
                  for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        dados = None
        for enc in ("utf-8-sig", "latin-1"):
            try:
                dados = open(caminho, encoding=enc).read()
                break
            except UnicodeDecodeError:
                continue
        if dados is None:
            sys.exit(f"Nao consegui ler {caminho}")
        try:
            sep = csv.Sniffer().sniff(dados[:8000], delimiters=";,\t").delimiter
        except csv.Error:
            sep = ";" if dados[:8000].count(";") > dados[:8000].count(",") else ","
        linhas = list(csv.reader(dados.splitlines(), delimiter=sep))

    # O cabecalho nem sempre e a primeira linha (planilha com titulo/logo em cima).
    melhor_i, melhor_n = 0, -1
    for i, l in enumerate(linhas[:15]):
        n = sum(1 for c in l if str(c).strip())
        if n > melhor_n:
            melhor_i, melhor_n = i, n
    cab = [norm(c) for c in linhas[melhor_i]]
    return [dict(zip(cab, [str(c).strip() for c in l]))
            for l in linhas[melhor_i + 1:] if any(str(c).strip() for c in l)], cab


def achar_coluna(cabecalho, pistas):
    for p in pistas:
        for col in cabecalho:
            if p.upper() in col:
                return col
    return None


# ---------------------------------------------------------------------------
# Analise
# ---------------------------------------------------------------------------

def idade_meses(aaaammdd):
    if not aaaammdd or len(aaaammdd) != 8 or not aaaammdd.isdigit():
        return None
    ano, mes = int(aaaammdd[:4]), int(aaaammdd[4:6])
    if not 1 <= mes <= 12:
        return None
    h = date.today()
    return (h.year - ano) * 12 + (h.month - mes)


def parse_valor(v):
    """Aceita 'R$ 45.000,00', '45000.00', '45,000.00'."""
    t = re.sub(r"[^\d,.\-]", "", str(v or ""))
    if not t:
        return 0.0
    if "," in t and "." in t:
        t = t.replace(".", "").replace(",", ".") if t.rfind(",") > t.rfind(".") \
            else t.replace(",", "")
    elif "," in t:
        t = t.replace(",", ".")
    try:
        return abs(float(t))
    except ValueError:
        return 0.0


def distribuicao(titulo, contador, total, unidade="clientes", faturamento=None):
    print(f"\n  {titulo}")
    if not contador:
        print("    (sem dados)")
        return
    for k, n in contador.most_common(12):
        pct = 100 * n / total if total else 0
        barra = "#" * max(1, round(pct / 4))
        linha = f"    {str(k)[:34]:<34} {n:>4} {unidade[:3]} {pct:>5.1f}%  {barra}"
        if faturamento and faturamento.get(k):
            linha += f"   R$ {faturamento[k]:>12,.0f}".replace(",", ".")
        print(linha)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("planilha", help="lista de clientes do cliente (.xlsx ou .csv)")
    ap.add_argument("--indice", help="indice regional (padrao: mais recente em saida/)")
    ap.add_argument("--corte", type=float, default=0.72,
                    help="similaridade minima para aceitar casamento por nome")
    args = ap.parse_args()

    # --- indice regional -----------------------------------------------------
    caminho_idx = args.indice
    if not caminho_idx:
        cands = sorted(glob.glob(os.path.join(OUT, "indice_regional_*.csv")))
        if not cands:
            sys.exit("Indice regional nao encontrado.\n"
                     "Rode antes:  python gerar_base.py --indice-amplo")
        caminho_idx = cands[-1]
    print(f"Indice  : {caminho_idx}")
    indice = list(csv.DictReader(open(caminho_idx, encoding="utf-8-sig")))
    print(f"          {len(indice):,} estabelecimentos ativos na regiao")

    por_cnpj = {so_digitos(r["cnpj"]): r for r in indice}
    por_chave = defaultdict(list)
    for r in indice:
        for campo in ("razao_social", "nome_fantasia"):
            k = chave(r.get(campo, ""))
            if k:
                por_chave[k].append(r)

    # indice invertido por token, para nao comparar contra tudo
    por_token = defaultdict(set)
    for i, r in enumerate(indice):
        for t in set(tokens(r.get("razao_social", "")) +
                     tokens(r.get("nome_fantasia", ""))):
            por_token[t].add(i)

    # --- planilha do cliente -------------------------------------------------
    clientes, cab = ler_planilha(args.planilha)
    col_cnpj = achar_coluna(cab, PISTAS["cnpj"])
    col_nome = achar_coluna(cab, PISTAS["nome"])
    col_valor = achar_coluna(cab, PISTAS["valor"])
    print(f"\nPlanilha: {args.planilha}")
    print(f"          {len(clientes):,} linhas | colunas detectadas -> "
          f"CNPJ={col_cnpj or '(nenhuma)'} | NOME={col_nome or '(nenhuma)'} | "
          f"VALOR={col_valor or '(nenhuma)'}")
    if not col_cnpj and not col_nome:
        sys.exit("Nenhuma coluna de CNPJ ou nome encontrada. Cabecalho lido: "
                 + ", ".join(cab))

    # --- casamento -----------------------------------------------------------
    casados, nao_encontrados = [], []
    vistos = set()
    for lin in clientes:
        bruto_nome = (lin.get(col_nome, "") if col_nome else "").strip()
        bruto_cnpj = so_digitos(lin.get(col_cnpj, "")) if col_cnpj else ""
        valor = parse_valor(lin.get(col_valor, "")) if col_valor else 0.0
        if not bruto_nome and not bruto_cnpj:
            continue

        alvo, conf, como = None, 0.0, ""
        if len(bruto_cnpj) == 14 and bruto_cnpj in por_cnpj:
            alvo, conf, como = por_cnpj[bruto_cnpj], 1.0, "CNPJ"
        elif bruto_nome:
            k = chave(bruto_nome)
            if k and k in por_chave:
                alvo, conf, como = por_chave[k][0], 0.99, "nome exato"
            elif k:
                cand = set()
                for t in tokens(bruto_nome):
                    cand |= por_token.get(t, set())
                melhor, melhor_s = None, 0.0
                for i in cand:
                    s = max(similaridade(bruto_nome, indice[i].get("razao_social", "")),
                            similaridade(bruto_nome, indice[i].get("nome_fantasia", "")))
                    if s > melhor_s:
                        melhor, melhor_s = indice[i], s
                if melhor and melhor_s >= args.corte:
                    alvo, conf, como = melhor, melhor_s, "nome aproximado"

        if alvo is None:
            nao_encontrados.append((bruto_nome or bruto_cnpj, valor))
            continue
        chave_dedupe = alvo["cnpj"]
        if chave_dedupe in vistos:
            # cliente recorrente: soma o faturamento no registro ja existente
            for c in casados:
                if c["cnpj"] == chave_dedupe:
                    c["valor"] += valor
                    c["compras"] += 1
                    break
            continue
        vistos.add(chave_dedupe)
        casados.append({**alvo, "entrada": bruto_nome or bruto_cnpj,
                        "confianca": conf, "como": como,
                        "valor": valor, "compras": 1})

    total = len(casados)
    print(f"\nCasamento: {total} de "
          f"{total + len(nao_encontrados)} clientes localizados na Receita "
          f"({100*total/max(total+len(nao_encontrados),1):.0f}%)")
    for como, n in Counter(c["como"] for c in casados).most_common():
        print(f"           {n:>4} por {como}")
    if nao_encontrados:
        print(f"           {len(nao_encontrados):>4} nao encontrados "
              f"(pessoa fisica, fora da regiao ou nome muito diferente)")

    if not total:
        sys.exit("\nNenhum cliente casado. Verifique se a coluna de nome esta certa "
                 "ou amplie ANEIS em config.py.")

    # --- perfil dos clientes reais -------------------------------------------
    fat = defaultdict(float)
    tem_valor = any(c["valor"] for c in casados)

    cnae_c, mun_c, porte_c, anel_c = Counter(), Counter(), Counter(), Counter()
    for c in casados:
        seg = C.CNAES.get(c["cnae_principal"], {}).get("segmento")
        rot = f"{c['cnae_principal']} {seg or '(fora do filtro)'}"
        cnae_c[rot] += 1
        fat[rot] += c["valor"]
        mun_c[c["municipio"] or "(?)"] += 1
        porte_c[c["porte"] or "(?)"] += 1
        anel_c[f"Anel {c['anel']}"] += 1

    print("\n" + "=" * 78)
    print("  PERFIL REAL DOS CLIENTES QUE JA COMPRAM")
    print("=" * 78)
    distribuicao("CNAE principal:", cnae_c, total,
                 faturamento=fat if tem_valor else None)
    distribuicao("Municipio:", mun_c, total)
    distribuicao("Anel de frete:", anel_c, total)
    distribuicao("Porte:", porte_c, total)

    caps = sorted(parse_valor(c["capital_social"]) for c in casados)
    if caps:
        def pct(p):
            return caps[min(int(len(caps) * p), len(caps) - 1)]
        print("\n  Capital social:")
        print(f"    minimo   R$ {caps[0]:>14,.0f}".replace(",", "."))
        print(f"    p10      R$ {pct(0.10):>14,.0f}".replace(",", "."))
        print(f"    mediana  R$ {pct(0.50):>14,.0f}".replace(",", "."))
        print(f"    p90      R$ {pct(0.90):>14,.0f}".replace(",", "."))
        print(f"    maximo   R$ {caps[-1]:>14,.0f}".replace(",", "."))

    idades = [i for i in (idade_meses(c["data_inicio"]) for c in casados) if i]
    if idades:
        idades.sort()
        print(f"\n  Idade da empresa: mediana {idades[len(idades)//2]//12} anos "
              f"| {sum(1 for i in idades if i < C.IDADE_MINIMA_MESES)} abaixo do "
              f"minimo atual de {C.IDADE_MINIMA_MESES} meses")

    # --- o filtro atual mataria quantos destes clientes? ---------------------
    print("\n" + "=" * 78)
    print("  TESTE DOS FILTROS ATUAIS CONTRA OS CLIENTES REAIS")
    print("=" * 78)

    cnaes_alvo = {c for c, m in C.CNAES.items() if m["nucleo"] or not C.SOMENTE_NUCLEO}
    motivos = defaultdict(list)
    for c in casados:
        cn = {c["cnae_principal"]} | {x.strip() for x in
                                      (c["cnaes_secundarios"] or "").split(",") if x.strip()}
        cap = parse_valor(c["capital_social"])
        if not (cn & cnaes_alvo):
            motivos["CNAE fora da lista"].append(c)
        if int(c["anel"]) > C.ANEL_MAXIMO:
            motivos[f"Municipio fora do anel {C.ANEL_MAXIMO}"].append(c)
        if cap > 0 and cap < C.CAPITAL_SOCIAL_MIN:
            motivos["Capital abaixo do minimo"].append(c)
        if cap > C.CAPITAL_SOCIAL_MAX:
            motivos["Capital acima do maximo"].append(c)
        im = idade_meses(c["data_inicio"])
        if im is not None and im < C.IDADE_MINIMA_MESES:
            motivos["Empresa nova demais"].append(c)

    perdidos = {c["cnpj"] for lst in motivos.values() for c in lst}
    if not perdidos:
        print("\n  Nenhum cliente atual seria descartado. Filtros validados.")
    else:
        fat_perdido = sum(c["valor"] for c in casados if c["cnpj"] in perdidos)
        fat_total = sum(c["valor"] for c in casados)
        print(f"\n  {len(perdidos)} de {total} clientes atuais "
              f"({100*len(perdidos)/total:.0f}%) seriam DESCARTADOS "
              f"pelo config.py atual.")
        if tem_valor and fat_total:
            print(f"  Isso representa R$ {fat_perdido:,.0f} de "
                  f"R$ {fat_total:,.0f} ({100*fat_perdido/fat_total:.0f}% do "
                  f"faturamento mapeado).".replace(",", "."))
        for motivo, lst in sorted(motivos.items(), key=lambda kv: -len(kv[1])):
            print(f"\n    {motivo}: {len(lst)}")
            for c in lst[:6]:
                extra = ""
                if "Capital" in motivo:
                    extra = f" | capital R$ {parse_valor(c['capital_social']):,.0f}".replace(",", ".")
                elif "CNAE" in motivo:
                    extra = f" | CNAE {c['cnae_principal']}"
                elif "anel" in motivo:
                    extra = f" | {c['municipio']} (anel {c['anel']})"
                print(f"      - {c['razao_social'][:44]}{extra}")
            if len(lst) > 6:
                print(f"      ... e mais {len(lst)-6}")

    # --- CNAEs frequentes que NAO estao no filtro ---------------------------
    fora = Counter()
    fat_fora = defaultdict(float)
    for c in casados:
        if c["cnae_principal"] not in C.CNAES:
            fora[c["cnae_principal"]] += 1
            fat_fora[c["cnae_principal"]] += c["valor"]
    if fora:
        print("\n" + "=" * 78)
        print("  CNAEs DE CLIENTES REAIS AUSENTES DO config.py")
        print("=" * 78)
        print("  Avalie incluir os de maior volume/faturamento em CNAES:\n")
        for cn, n in fora.most_common(10):
            linha = f"    {cn}  {n:>3} clientes"
            if tem_valor and fat_fora[cn]:
                linha += f"   R$ {fat_fora[cn]:>12,.0f}".replace(",", ".")
            print(linha)

    # --- saida ---------------------------------------------------------------
    destino = os.path.join(OUT, "icp_clientes_casados.csv")
    cols = ["entrada", "como", "confianca", "cnpj", "razao_social", "nome_fantasia",
            "municipio", "anel", "cnae_principal", "capital_social", "porte",
            "data_inicio", "telefone_1", "email", "compras", "valor"]
    with open(destino, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for c in sorted(casados, key=lambda x: -x["valor"]):
            w.writerow({**c, "confianca": f"{c['confianca']:.2f}"})

    if nao_encontrados:
        alvo2 = os.path.join(OUT, "icp_nao_encontrados.csv")
        with open(alvo2, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["entrada", "valor"])
            w.writerows(nao_encontrados)
        print(f"\n  Nao encontrados : {alvo2}")

    print(f"  Clientes casados: {destino}")
    revisar = [c for c in casados if c["como"] == "nome aproximado"]
    if revisar:
        print(f"\n  ATENCAO: {len(revisar)} casamentos por aproximacao. "
              f"Confira a coluna 'confianca' antes de decidir com base neles.")


if __name__ == "__main__":
    main()
