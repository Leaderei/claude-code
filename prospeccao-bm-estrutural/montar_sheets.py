#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Monta a planilha de trabalho do SDR (.xlsx) a partir da base gerada.

Sem argumento: gera a planilha VAZIA, so com a estrutura, para subir ao
Google Sheets antes da base existir.
Com o CSV: preenche as linhas.

    python montar_sheets.py                                   # so estrutura
    python montar_sheets.py saida/base_bm_estrutural_2026-08.csv
"""

import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Paleta oficial Leaderei. Planilha de trabalho usa o modo claro: a base e
# lida e impressa o dia inteiro pelas SDRs.
LARANJA = "DF4E01"
PRETO = "000000"
CINZA_2 = "313131"
CINZA_1 = "606060"
CINZA_CLARO = "F7F7F7"
BORDA = "E0E0E0"

# Ordem identica a saida do gerar_base.py: a importacao do CSV cai 1:1.
# As duas ultimas sao de uso manual da SDR.
COLUNAS = [
    ("cnpj", 20), ("razao_social", 38), ("nome_fantasia", 24), ("segmento", 16),
    ("prioridade", 11), ("score", 8), ("anel", 7), ("municipio", 18),
    ("bairro", 18), ("endereco", 30), ("cep", 11), ("telefone_1", 17),
    ("telefone_2", 17), ("email", 30), ("dominio", 24), ("site", 28),
    ("capital_social", 15), ("porte", 9), ("matriz", 9), ("cnae_principal", 15),
    ("instagram", 22), ("responsavel", 14), ("tipologia_obra", 20),
    ("status_sdr", 24), ("obs", 40),
    ("data_ultimo_contato", 18), ("proximo_passo", 26),
]

# Espelha a cadencia multicanal de 13 dias da Leaderei.
STATUS = [
    "Nao iniciado", "D1 LinkedIn", "D3 WhatsApp", "D5 E-mail", "D7 Ligacao",
    "D9 Ligacao", "D11 Ligacao", "D13 Breakup", "Conectado",
    "Reuniao agendada", "Visita agendada", "Orcamento enviado",
    "Ganho", "Perdido", "Fora do ICP",
]
TIPOLOGIA = [
    "Casa terrea", "Sobrado", "Condominio residencial", "Predio ate 4 pav",
    "Predio alto", "Comercial", "Industrial", "Nao identificado",
]
RESPONSAVEIS = ["Fernanda", "Bruna", "Luciana"]

INSTRUCOES = [
    ("titulo", "Base de prospeccao — BM Estrutural"),
    ("sub", "Engenharias, construtoras e escritorios de arquitetura na macro-regiao de Louveira"),
    ("", ""),
    ("h", "Como a fila funciona"),
    ("p", "A aba Base ja vem ordenada por score. Trabalhe de cima para baixo."),
    ("p", "Prioridade A (score 60+) — fila principal do SDR."),
    ("p", "Prioridade B (40-59) — trabalhar quando a fila A estiver vazia."),
    ("p", "Prioridade C (abaixo de 40) — so automacao e e-mail, sem ligacao."),
    ("", ""),
    ("h", "Divisao por anel de frete"),
    ("p", "Anel 1 (ate 30 km: Louveira, Vinhedo, Jundiai, Itupeva, Valinhos, Itatiba, Cabreuva)"),
    ("p", "   -> qualificar e passar para o vendedor externo da regiao."),
    ("p", "Anel 2 (30-60 km: Campinas, Indaiatuba, Atibaia, Salto e demais)"),
    ("p", "   -> sem vendedor externo. Qualificar e agendar reuniao remota."),
    ("", ""),
    ("h", "Coluna status_sdr — cadencia de 13 dias"),
    ("p", "Os status D1 a D13 seguem a cadencia multicanal padrao."),
    ("p", "Referencia: 6 a 12 tentativas por lead, alternando canais."),
    ("p", "Abaixo de 6 tentativas a conversao cai de forma relevante."),
    ("p", "Nao marque Perdido antes do D13 Breakup."),
    ("", ""),
    ("h", "Coluna tipologia_obra — a mais importante"),
    ("p", "E o que separa lead bom de ruido. Engenharia que faz casa e obra"),
    ("p", "de pequeno porte e o alvo; quem so faz predio alto ou galpao nao e."),
    ("p", "Preencha assim que descobrir, na ligacao ou olhando o site."),
    ("", ""),
    ("h", "Colunas site e dominio"),
    ("p", "Vem do e-mail registrado na Receita Federal, quando o dominio e proprio."),
    ("p", "Vazio nao significa que a empresa nao tem site — significa que ela usa"),
    ("p", "e-mail generico (gmail, hotmail). Nesses casos procure no Google."),
    ("p", "Escritorio de arquitetura costuma ter Instagram e nao ter site."),
    ("", ""),
    ("h", "Origem dos dados"),
    ("p", "Dados Abertos do CNPJ — Receita Federal. Atualizacao mensal."),
    ("p", "Filtros: empresa ativa, aberta ha 24+ meses, nao-MEI,"),
    ("p", "capital social entre R$ 20 mil e R$ 5 milhoes."),
    ("p", "O teto de capital exclui incorporadora de larga escala de proposito."),
    ("", ""),
    ("h", "LGPD"),
    ("p", "Dados cadastrais de PJ sao publicos. Todo disparo precisa de opt-out."),
    ("p", "Registre pedidos de descadastro na coluna obs e marque Fora do ICP."),
]


def aba_base(wb, linhas):
    ws = wb.create_sheet("Base")
    fino = Side(style="thin", color=BORDA)

    for i, (nome, larg) in enumerate(COLUNAS, start=1):
        c = ws.cell(row=1, column=i, value=nome)
        c.fill = PatternFill("solid", fgColor=LARANJA)
        c.font = Font(name="Manrope", bold=True, size=10, color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=fino)
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[1].height = 26

    campos = [n for n, _ in COLUNAS]
    for r, linha in enumerate(linhas, start=2):
        for i, nome in enumerate(campos, start=1):
            c = ws.cell(row=r, column=i, value=linha.get(nome, ""))
            c.font = Font(name="Manrope", size=10, color=CINZA_2)
            c.border = Border(bottom=Border(bottom=fino).bottom)

    ws.freeze_panes = "C2"
    ultima = max(len(linhas) + 1, 2)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{ultima}"

    # Listas suspensas — ate a linha 5000 para cobrir crescimento da base.
    fim = max(ultima, 5000)
    for col_nome, valores in (("status_sdr", STATUS), ("tipologia_obra", TIPOLOGIA),
                              ("responsavel", RESPONSAVEIS)):
        idx = campos.index(col_nome) + 1
        dv = DataValidation(type="list", formula1='"' + ",".join(valores) + '"',
                            allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        letra = get_column_letter(idx)
        dv.add(f"{letra}2:{letra}{fim}")

    # Prioridade destacada — laranja em A, cinza no resto.
    letra_p = get_column_letter(campos.index("prioridade") + 1)
    faixa = f"{letra_p}2:{letra_p}{fim}"
    ws.conditional_formatting.add(faixa, CellIsRule(
        operator="equal", formula=['"A"'],
        fill=PatternFill("solid", bgColor="FDE3D4"),
        font=Font(name="Manrope", bold=True, color=LARANJA)))
    ws.conditional_formatting.add(faixa, CellIsRule(
        operator="equal", formula=['"C"'],
        font=Font(name="Manrope", color=CINZA_1)))
    return ws


def aba_instrucoes(wb):
    ws = wb.create_sheet("Como usar")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 96
    ws.sheet_view.showGridLines = False

    caminho_logo = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "logo_leaderei.png")
    if os.path.exists(caminho_logo):
        try:
            from openpyxl.drawing.image import Image
            img = Image(caminho_logo)
            img.height, img.width = 34, int(34 * 292.86 / 64.81)
            ws.add_image(img, "B2")
        except Exception:
            pass

    r = 5
    for tipo, texto in INSTRUCOES:
        c = ws.cell(row=r, column=2, value=texto)
        if tipo == "titulo":
            c.font = Font(name="Manrope", bold=True, size=18, color=PRETO)
            ws.row_dimensions[r].height = 26
        elif tipo == "sub":
            c.font = Font(name="Manrope", size=11, color=CINZA_1)
        elif tipo == "h":
            c.font = Font(name="Manrope", bold=True, size=12, color=LARANJA)
            ws.row_dimensions[r].height = 22
        else:
            c.font = Font(name="Manrope", size=10.5, color=CINZA_2)
        c.alignment = Alignment(vertical="center")
        r += 1
    return ws


def main():
    caminho_csv = sys.argv[1] if len(sys.argv) > 1 else None
    linhas = []
    if caminho_csv:
        if not os.path.exists(caminho_csv):
            sys.exit(f"CSV nao encontrado: {caminho_csv}")
        linhas = list(csv.DictReader(open(caminho_csv, encoding="utf-8-sig")))

    wb = Workbook()
    wb.remove(wb.active)
    aba_instrucoes(wb)
    aba_base(wb, linhas)
    wb.active = 1

    destino = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "saida", "base_bm_estrutural.xlsx")
    os.makedirs(os.path.dirname(destino), exist_ok=True)
    wb.save(destino)
    print(f"Planilha gerada: {destino}")
    print(f"  {len(linhas):,} linhas de dados")
    print("  Suba no Google Drive ou importe no Google Sheets.")


if __name__ == "__main__":
    main()
